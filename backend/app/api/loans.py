from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from typing import List, Optional
from uuid import UUID
from app.database import get_db
from app.schemas.loan import LoanCreate, LoanUpdate, LoanResponse, LoanSummary
from app.schemas.payment import PaymentResponse
from app.schemas.alert import AlertResponse
from app.schemas.attachment import AttachmentResponse
from app.services import loan_service
from app.services.interest_service import accrue_interest, recalculate_loan
from app.services import attachment_service, payment_service
from app.models.attachment import AttachmentParent
from app.models.alert import Alert
from fastapi.responses import StreamingResponse
import io
import openpyxl
from fpdf import FPDF

router = APIRouter()

@router.get("/", response_model=List[dict])
def list_loans(
    direction:  Optional[str]  = Query(None),
    status:     Optional[str]  = Query(None),
    person_id:  Optional[UUID] = Query(None),
    currency:   Optional[str]  = Query(None),
    db: Session = Depends(get_db)
):
    return loan_service.get_all(db, direction, status, person_id, currency)

@router.get("/overdue", response_model=List[dict])
def list_overdue(db: Session = Depends(get_db)):
    return loan_service.get_overdue(db)

@router.get("/due-soon", response_model=List[dict])
def list_due_soon(
    days: int = Query(7),
    db: Session = Depends(get_db)
):
    return loan_service.get_due_soon(db, days)

@router.get("/{loan_id}", response_model=dict)
def get_loan(loan_id: UUID, db: Session = Depends(get_db)):
    result = loan_service.get_summary_by_id(db, loan_id)
    if not result:
        raise HTTPException(status_code=404, detail="Loan not found")
    return result

@router.post("/", response_model=LoanResponse, status_code=201)
def create_loan(data: LoanCreate, db: Session = Depends(get_db)):
    loan = loan_service.create(db, data)
    # loan_service automatically runs the new math engine, so we just return!
    return loan_service.get_summary_by_id(db, loan.id) or loan

@router.patch("/{loan_id}", response_model=LoanResponse)
def update_loan(loan_id: UUID, data: LoanUpdate, db: Session = Depends(get_db)):
    result = loan_service.update(db, loan_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="Loan not found")

    # --- THE FIX: Save the calculated diff into the Audit Log ---
    diff_text = getattr(result, "_audit_diff", None)
    if diff_text:
        from app.models.audit_log import AuditLog

        audit_entry = AuditLog(
            loan_id=loan_id,
            action="Updated Details",
            description=diff_text
        )
        db.add(audit_entry)
        db.commit()
    # ------------------------------------------------------------

    return result

@router.delete("/{loan_id}", status_code=204)
def delete_loan(loan_id: UUID, db: Session = Depends(get_db)):
    success = loan_service.delete(db, loan_id)
    if not success:
        raise HTTPException(status_code=404, detail="Loan not found")

@router.post("/{loan_id}/cancel", response_model=LoanResponse)
def cancel_loan(loan_id: UUID, db: Session = Depends(get_db)):
    result = loan_service.cancel(db, loan_id)
    if not result:
        raise HTTPException(status_code=404, detail="Loan not found")

    # --- THE FIX: Save the cancellation to the Audit Log ---
    from app.models.audit_log import AuditLog
    audit_entry = AuditLog(
        loan_id=loan_id,
        action="Cancelled",
        description="Loan was manually cancelled. Remaining balance wiped to $0."
    )
    db.add(audit_entry)
    db.commit()
    # -------------------------------------------------------

    return result

@router.get("/{loan_id}/payments", response_model=List[PaymentResponse])
def get_loan_payments(loan_id: UUID, db: Session = Depends(get_db)):
    return payment_service.get_by_loan(db, loan_id)

@router.get("/{loan_id}/attachments", response_model=List[AttachmentResponse])
def get_loan_attachments(loan_id: UUID, db: Session = Depends(get_db)):
    return attachment_service.get_by_parent(db, loan_id, AttachmentParent.loan)

@router.get("/{loan_id}/alerts", response_model=List[AlertResponse])
def get_loan_alerts(loan_id: UUID, db: Session = Depends(get_db)):
    loan = loan_service.get_by_id(db, loan_id)
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    return db.query(Alert).filter(
        Alert.loan_id == loan_id,
        Alert.is_dismissed == False
    ).order_by(Alert.trigger_date.desc()).all()

@router.get("/{loan_id}/interest", response_model=List[dict])
def get_loan_interest(loan_id: UUID, db: Session = Depends(get_db)):
    from sqlalchemy import text
    rows = db.execute(text(
        "SELECT period_start, period_end, opening_balance, "
        "interest_accrued, closing_balance, calc_type, rate_applied "
        "FROM interest_ledger WHERE loan_id = :loan_id ORDER BY period_start"
    ), {"loan_id": str(loan_id)}).mappings().all()
    return [
        {
            "period_start":     str(e["period_start"]),
            "period_end":       str(e["period_end"]),
            "opening_balance":  float(e["opening_balance"]),
            "interest_accrued": float(e["interest_accrued"]),
            "closing_balance":  float(e["closing_balance"]),
            "calc_type":        str(e["calc_type"]),
            "rate_applied":     float(e["rate_applied"]),
        }
        for e in rows
    ]

@router.get("/{loan_id}/audit", response_model=List[dict])
def get_loan_audit(loan_id: UUID, db: Session = Depends(get_db)):
    from app.models.audit_log import AuditLog
    logs = db.query(AuditLog).filter(
        AuditLog.loan_id == loan_id
    ).order_by(AuditLog.changed_at.desc()).all()

    clean_logs = []
    for l in logs:
        # THE FIX: Skip the automatic database ghost rows that have no details!
        if (l.action == 'updated' or l.action == 'Updated') and not l.description:
            continue

        clean_logs.append({
            "action":        l.action,
            "changed_field": l.changed_field,
            "old_value":     l.old_value,
            "new_value":     l.new_value,
            "description":   l.description,
            "changed_at":    l.changed_at,
        })

    return clean_logs

@router.get("/{loan_id}/statement/download")
def download_statement(loan_id: UUID, format: str = "txt", db: Session = Depends(get_db)):
    # 1. Fetch Enriched Loan Summary & Payments
    loan = loan_service.get_summary_by_id(db, loan_id)
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")

    payments = payment_service.get_by_loan(db, loan_id)

    # 2. Log the export directly to the Audit Log!
    from app.models.audit_log import AuditLog
    audit_entry = AuditLog(
        loan_id=loan_id,
        action="Statement Exported",
        description=f"User downloaded the loan statement in {format.upper()} format."
    )
    db.add(audit_entry)
    db.commit()

    # --- NEW: 3. Prepare the Enriched Loan Summary Header ---
    currency = loan.get("currency", "USD")
    rate = loan.get("interest_rate", 0)
    period = loan.get("interest_period", "monthly")
    rate_str = f"{rate}% ({period})" if rate and float(rate) > 0 else "0%"

    direction = loan.get("direction", "lent")
    party_role = "Borrower" if direction == "lent" else "Lender"
    direction_text = "Money Lent (They owe me)" if direction == "lent" else "Money Borrowed (I owe them)"

    # --- NEW FIX: Fetch fees to include their GST in the totals ---
    from app.services import loan_fee_service
    fees = loan_fee_service.get_by_loan(db, loan_id)
    # --------------------------------------------------------------

    # --- NEW: Calculate exact totals for the export ---
    total_principal_paid = sum((p.principal_component or 0) for p in payments)
    total_interest_paid = sum((p.interest_component or 0) for p in payments)
    total_tax_paid = 0

    # 1.Tax from payments
    for p in payments:
        if p.is_manual:
            total_tax_paid += (p.tax_amount or 0)
        else:
            p_rate = p.tax_rate or 0
            if p_rate > 0:
                total_tax_paid += (p.amount * (p_rate / 100))
    # --------------------------------------------------
    # 2. Tax from fees (The Bug Fix!)
    for f in fees:
        f_rate = f.tax_rate or 0
        if f_rate > 0:
            total_tax_paid += (f.amount * (f_rate / 100))
    # --------------------------------------------------

    summary_data = [
        ["Loan ID:", str(loan.get("id", loan_id))],
        [f"Party ({party_role}):", str(loan.get("person_name", "Unknown"))],
        ["Direction:", direction_text],
        ["Date Issued:", str(loan.get("date_issued", "N/A"))],
        ["Purpose:", str(loan.get("purpose") or "N/A")],
        ["Original Principal:", f"{currency} {float(loan.get('principal', 0)):.2f}"],
        ["Interest Rate:", rate_str],

        # --- NEW: Inject into the statement header ---
        ["Total Principal Paid:", f"{currency} {total_principal_paid:.2f}"],
        ["Total Interest Paid:", f"{currency} {total_interest_paid:.2f}"],
        ["Total Tax/GST Paid:", f"{currency} {total_tax_paid:.2f}"],
        # ---------------------------------------------

        ["Total Cash Paid to Date:", f"{currency} {float(loan.get('total_paid', 0)):.2f}"],
        ["Current Balance Due:", f"{currency} {float(loan.get('balance_due', 0)):.2f}"],
        ["Status:", str(loan.get("status", "unknown")).upper()]
    ]
    # --------------------------------------------------------

    # 4. Prepare the Payment Data Rows
    headers = ["Date", "Description", "Amount", "Principal Paid", "Interest Paid"]
    rows = []
    for p in payments:
        rows.append([
            str(p.payment_date),
            "Payment Received",
            f"{currency} {p.amount:.2f}",
            f"{currency} {p.principal_component or 0:.2f}",
            f"{currency} {p.interest_component or 0:.2f}"
        ])

    # 5. Generate the requested file type
    if format == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Loan Statement"

        # Write Summary Header
        ws.append(["OFFICIAL LOAN STATEMENT"])
        ws.append([]) # Empty row
        for item in summary_data:
            ws.append(item)

        ws.append([]) # Empty row
        ws.append(["--- TRANSACTION HISTORY ---"])
        ws.append(headers)
        for r in rows:
            ws.append(r)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=loan_statement.xlsx"}
        )

    elif format == "pdf":
        pdf = FPDF()
        pdf.add_page()

        # Write Summary Header
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, txt="OFFICIAL LOAN STATEMENT", ln=True, align='C')
        pdf.ln(5)

        pdf.set_font("Arial", '', 11)
        for label, val in summary_data:
            pdf.set_font("Arial", 'B', 11)
            pdf.cell(45, 8, txt=label)
            pdf.set_font("Arial", '', 11)
            pdf.cell(0, 8, txt=val, ln=True)

        pdf.ln(10)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, txt="--- TRANSACTION HISTORY ---", ln=True)

        # Draw Table Headers
        pdf.set_font("Arial", 'B', 10)
        col_widths = [25, 45, 30, 40, 40]
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, str(header), border=1)
        pdf.ln()

        # Draw Table Rows
        pdf.set_font("Arial", '', 10)
        for r in rows:
            for i, item in enumerate(r):
                pdf.cell(col_widths[i], 10, str(item), border=1)
            pdf.ln()

        pdf_bytes = pdf.output(dest='S')
        stream = io.BytesIO(pdf_bytes)
        return StreamingResponse(
            stream,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=loan_statement.pdf"}
        )

    else: # Default TXT
        stream = io.StringIO()
        stream.write("=========================================\n")
        stream.write("        OFFICIAL LOAN STATEMENT          \n")
        stream.write("=========================================\n\n")

        for label, val in summary_data:
            stream.write(f"{label:<25} {val}\n")

        stream.write("\n-----------------------------------------------------------------\n")
        stream.write(f"{' | '.join(headers)}\n")
        stream.write("-----------------------------------------------------------------\n")
        for r in rows:
            stream.write(f"{' | '.join(r)}\n")

        byte_stream = io.BytesIO(stream.getvalue().encode('utf-8'))
        return StreamingResponse(
            byte_stream,
            media_type="text/plain",
            headers={"Content-Disposition": f"attachment; filename=loan_statement.txt"}
        )
